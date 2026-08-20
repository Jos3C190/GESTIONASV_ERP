"""Use cases for deterministic warehouse locations and staged bulk creation."""

from __future__ import annotations

import csv
import hashlib
import io
import itertools
import json
import re
import unicodedata
import uuid
import zipfile
from collections.abc import Mapping, Sequence
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any, NoReturn

from app.core.exceptions import ConflictError, ValidationError
from app.domain.entities.location import (
    LOCATION_COMPONENT_KEYS,
    LOCATION_LIFECYCLE_STATUSES,
    LOCATION_TYPES,
    CodeProjection,
    CodeSegment,
    LocationBatchRecord,
    LocationCodeScheme,
    LocationRecord,
    WarehouseLocationScope,
    normalize_location_component,
    project_location_code,
)
from app.domain.entities.warehouse_capacity import (
    CAPACITY_ENFORCEMENT_MODES,
    CAPACITY_PROFILES,
    PhysicalCapacity,
)
from app.domain.ports.location_repository import LocationRepository

MAX_BATCH_ROWS = 50_000
MAX_DECIMAL_PLACES = 6
MAX_IMPORT_BYTES = 20 * 1024 * 1024
MAX_XLSX_ENTRIES = 10_000
MAX_XLSX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 1_000
_PREFIXED_NUMBER = re.compile(r"^(.*?)(\d+)$")

_HEADER_ALIASES = {
    "area": "area",
    "zona": "area",
    "aisle": "aisle",
    "pasillo": "aisle",
    "rack": "rack",
    "estante": "rack",
    "level": "level",
    "nivel": "level",
    "position": "position",
    "posicion": "position",
    "bin": "position",
    "certified_max_weight_kg": "certified_max_weight_kg",
    "peso_maximo_certificado_kg": "certified_max_weight_kg",
    "operational_max_weight_kg": "operational_max_weight_kg",
    "peso_maximo_operativo_kg": "operational_max_weight_kg",
    "certified_usable_volume_m3": "certified_usable_volume_m3",
    "volumen_util_certificado_m3": "certified_usable_volume_m3",
    "operational_usable_volume_m3": "operational_usable_volume_m3",
    "volumen_util_operativo_m3": "operational_usable_volume_m3",
    "capacity_profile": "capacity_profile",
    "perfil_capacidad": "capacity_profile",
    "capacity_enforcement_mode": "capacity_enforcement_mode",
    "modo_control_capacidad": "capacity_enforcement_mode",
    "storage_eligible": "storage_eligible",
    "apta_para_almacenamiento": "storage_eligible",
    "usable_length_m": "usable_length_m",
    "largo_util_m": "usable_length_m",
    "usable_width_m": "usable_width_m",
    "ancho_util_m": "usable_width_m",
    "usable_height_m": "usable_height_m",
    "alto_util_m": "usable_height_m",
    "capacity_group_id": "capacity_group_id",
    "grupo_capacidad_id": "capacity_group_id",
    "notes": "notes",
    "notas": "notes",
    "location_type": "location_type",
    "tipo": "location_type",
    "lifecycle_status": "lifecycle_status",
    "estado": "lifecycle_status",
    "barcode": "barcode",
    "codigo_de_barras": "barcode",
    "verification_code": "verification_code",
    "codigo_de_verificacion": "verification_code",
    "pick_sequence": "pick_sequence",
    "secuencia_de_picking": "pick_sequence",
    "putaway_sequence": "putaway_sequence",
    "secuencia_de_ubicacion": "putaway_sequence",
    "external_id": "external_id",
    "id_externo": "external_id",
}
_LOCATION_IMPORT_PATCH_FIELDS = frozenset(_HEADER_ALIASES.values())


def _nfkc_text(value: object, *, upper: bool = False) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = unicodedata.normalize("NFKC", str(value)).strip()
    return text.upper() if upper else text


def _header(value: object) -> str:
    text = _nfkc_text(value).casefold()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def _canonical_checksum(value: object) -> str:
    encoded = json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _strict_int(value: object, *, message: str, code: str) -> int:
    if isinstance(value, bool):
        raise ValidationError(message, code=code)
    normalized = _nfkc_text(value)
    if not re.fullmatch(r"[+-]?\d+", normalized):
        raise ValidationError(message, code=code)
    return int(normalized)


def _strict_decimal(value: object, *, message: str, code: str) -> Decimal:
    if isinstance(value, bool) or value in (None, ""):
        raise ValidationError(message, code=code)
    try:
        parsed = Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValidationError(message, code=code) from exc
    if not parsed.is_finite() or parsed.as_tuple().exponent < -MAX_DECIMAL_PLACES:
        raise ValidationError(message, code=code)
    return parsed


def _strict_bool(value: object, *, message: str, code: str) -> bool:
    if isinstance(value, bool):
        return value
    normalized = _nfkc_text(value).casefold()
    if normalized in {"1", "true", "si", "sí", "yes"}:
        return True
    if normalized in {"0", "false", "no"}:
        return False
    raise ValidationError(message, code=code)


def _fail_validation(message: str, code: str) -> NoReturn:
    raise ValidationError(message, code=code)


def _validate_xlsx_archive(content: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_XLSX_ENTRIES:
                _fail_validation(
                    "El XLSX contiene demasiados archivos internos.",
                    "location_import_xlsx_too_large",
                )
            if any(
                entry.flag_bits & 0x1
                or entry.filename.startswith(("/", "\\"))
                or ".." in entry.filename.replace("\\", "/").split("/")
                for entry in entries
            ):
                _fail_validation(
                    "El XLSX contiene entradas cifradas o rutas internas no permitidas.",
                    "location_import_xlsx_invalid",
                )
            uncompressed = sum(entry.file_size for entry in entries)
            compressed = sum(max(entry.compress_size, 1) for entry in entries)
            if (
                uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES
                or (
                    uncompressed > MAX_IMPORT_BYTES
                    and uncompressed / compressed > MAX_XLSX_COMPRESSION_RATIO
                )
                or any(
                    entry.file_size > MAX_IMPORT_BYTES
                    and entry.file_size / max(entry.compress_size, 1) > MAX_XLSX_COMPRESSION_RATIO
                    for entry in entries
                )
            ):
                _fail_validation(
                    "El contenido descomprimido del XLSX excede el límite seguro.",
                    "location_import_xlsx_too_large",
                )
    except zipfile.BadZipFile as exc:
        raise ValidationError(
            "El archivo XLSX no es válido.", code="location_import_xlsx_invalid"
        ) from exc


def _normalize_operational_values(values: Mapping[str, Any]) -> dict[str, Any]:  # noqa: C901
    normalized = dict(values)

    raw_location_type = values.get("location_type", "standard")
    location_type = _nfkc_text(raw_location_type).casefold()
    if location_type not in LOCATION_TYPES:
        raise ValidationError("El tipo de ubicación no es válido.", code="location_type_invalid")
    normalized["location_type"] = location_type

    for name, code in (
        ("certified_max_weight_kg", "location_certified_weight_invalid"),
        ("operational_max_weight_kg", "location_operational_weight_invalid"),
        ("certified_usable_volume_m3", "location_certified_volume_invalid"),
        ("operational_usable_volume_m3", "location_operational_volume_invalid"),
        ("usable_length_m", "location_usable_dimensions_invalid"),
        ("usable_width_m", "location_usable_dimensions_invalid"),
        ("usable_height_m", "location_usable_dimensions_invalid"),
    ):
        raw = values.get(name)
        normalized[name] = (
            None
            if raw in (None, "")
            else _strict_decimal(
                raw,
                message=f"{name} debe ser un número positivo con hasta seis decimales.",
                code=code,
            )
        )
        if normalized[name] is not None and normalized[name] <= 0:
            raise ValidationError(f"{name} debe ser mayor que cero.", code=code)
    profile = _nfkc_text(values.get("capacity_profile", "general_mixed")).casefold()
    if profile not in CAPACITY_PROFILES:
        raise ValidationError(
            "El perfil de capacidad no es válido.", code="capacity_profile_invalid"
        )
    normalized["capacity_profile"] = profile
    enforcement = _nfkc_text(values.get("capacity_enforcement_mode", "disabled")).casefold()
    if enforcement not in CAPACITY_ENFORCEMENT_MODES:
        raise ValidationError(
            "El modo de control de capacidad no es válido.",
            code="capacity_enforcement_mode_invalid",
        )
    normalized["capacity_enforcement_mode"] = enforcement
    raw_eligible = values.get("storage_eligible")
    normalized["storage_eligible"] = (
        location_type not in {"receiving", "quality", "packing", "shipping", "virtual"}
        if raw_eligible in (None, "")
        else _strict_bool(
            raw_eligible,
            message="storage_eligible debe ser verdadero o falso.",
            code="location_storage_eligible_invalid",
        )
    )
    raw_group_id = values.get("capacity_group_id")
    if raw_group_id in (None, ""):
        normalized["capacity_group_id"] = None
    else:
        try:
            normalized["capacity_group_id"] = uuid.UUID(str(raw_group_id))
        except (TypeError, ValueError, AttributeError) as exc:
            raise ValidationError(
                "El grupo de capacidad no es válido.",
                code="location_capacity_group_invalid",
            ) from exc

    try:
        PhysicalCapacity(
            certified_max_weight_kg=normalized["certified_max_weight_kg"],
            operational_max_weight_kg=normalized["operational_max_weight_kg"],
            certified_usable_volume_m3=normalized["certified_usable_volume_m3"],
            operational_usable_volume_m3=normalized["operational_usable_volume_m3"],
            capacity_profile=profile,
            capacity_enforcement_mode=enforcement,
            storage_eligible=normalized["storage_eligible"],
            usable_length_m=normalized["usable_length_m"],
            usable_width_m=normalized["usable_width_m"],
            usable_height_m=normalized["usable_height_m"],
        )
    except ValueError as exc:
        raise ValidationError(str(exc), code="location_capacity_configuration_invalid") from exc

    raw_lifecycle_status = values.get("lifecycle_status", "active")
    lifecycle_status = _nfkc_text(raw_lifecycle_status).casefold()
    if lifecycle_status not in LOCATION_LIFECYCLE_STATUSES:
        raise ValidationError(
            "El estado de la ubicación no es válido.", code="location_status_invalid"
        )
    normalized["lifecycle_status"] = lifecycle_status
    normalized["is_active"] = lifecycle_status != "retired"

    for name in ("pick_sequence", "putaway_sequence"):
        raw = values.get(name)
        if raw in (None, ""):
            normalized[name] = None
            continue
        parsed = _strict_int(
            raw,
            message=f"{name} debe ser un entero no negativo.",
            code="location_sequence_invalid",
        )
        if parsed < 0:
            raise ValidationError(
                f"{name} debe ser un entero no negativo.",
                code="location_sequence_invalid",
            )
        normalized[name] = parsed

    for name in ("barcode", "verification_code", "external_id", "notes"):
        text = _nfkc_text(values.get(name))
        normalized[name] = text or None
    return normalized


def _present_import_fields(values: Mapping[str, Any]) -> list[str]:
    return [field for field in _LOCATION_IMPORT_PATCH_FIELDS if field in values]


def _normalize_physical_values(
    scheme: LocationCodeScheme, values: Mapping[str, Any]
) -> tuple[dict[str, Any], Any]:
    projection = project_location_code(scheme, values)
    normalized = _normalize_operational_values(values)
    by_key = {segment.key: segment for segment in scheme.segments}
    for key in LOCATION_COMPONENT_KEYS:
        segment = by_key.get(key) or CodeSegment(
            key=key,
            label=key.capitalize(),
            width=0,
            required=key != "area",
        )
        normalized[key] = normalize_location_component(values.get(key), segment) or None
    projection = CodeProjection(
        code=projection.code,
        normalized_components={
            key: str(normalized.get(key) or "") for key in LOCATION_COMPONENT_KEYS
        },
        scheme_id=projection.scheme_id,
        scheme_version=projection.scheme_version,
    )
    normalized["scheme_id"] = projection.scheme_id
    normalized["scheme_version"] = projection.scheme_version
    return normalized, projection


def _guard_axis_cardinality(first: int, last: int, step: int) -> None:
    if first > last:
        raise ValidationError(
            "El inicio de un eje no puede ser mayor que su final.",
            code="location_axis_inverted",
        )
    if ((last - first) // step) + 1 > MAX_BATCH_ROWS:
        raise ValidationError(
            f"Un eje no puede superar {MAX_BATCH_ROWS} valores.",
            code="location_batch_too_large",
        )


def _expand_explicit_axis(values: Sequence[object]) -> list[str]:
    if len(values) > MAX_BATCH_ROWS:
        raise ValidationError(
            f"Un eje no puede superar {MAX_BATCH_ROWS} valores.",
            code="location_batch_too_large",
        )
    result = [_nfkc_text(value, upper=True) for value in values]
    if any(not value for value in result):
        raise ValidationError(
            "Los valores de un eje no pueden estar vacíos.",
            code="location_axis_value_empty",
        )
    if len(result) != len(set(result)):
        raise ValidationError(
            "Un eje no puede repetir valores.", code="location_axis_value_repeated"
        )
    return result


def _expand_axis(axis: Mapping[str, Any]) -> list[str]:
    values = axis.get("values")
    if values:
        return _expand_explicit_axis(values)

    start = _nfkc_text(axis.get("start"), upper=True)
    end = _nfkc_text(axis.get("end"), upper=True)
    try:
        step = int(axis.get("step") or 1)
    except (TypeError, ValueError) as exc:
        raise ValidationError(
            "El incremento del eje debe ser un entero positivo.",
            code="location_axis_invalid",
        ) from exc
    if not start or not end or step < 1:
        raise ValidationError(
            "Cada eje debe incluir valores o un rango válido.", code="location_axis_invalid"
        )
    start_match = _PREFIXED_NUMBER.match(start)
    end_match = _PREFIXED_NUMBER.match(end)
    if start_match and end_match and start_match.group(1) == end_match.group(1):
        prefix = start_match.group(1)
        first, last = int(start_match.group(2)), int(end_match.group(2))
        width = max(len(start_match.group(2)), len(end_match.group(2)))
        _guard_axis_cardinality(first, last, step)
        return [f"{prefix}{number:0{width}d}" for number in range(first, last + 1, step)]
    if len(start) == len(end) == 1 and start.isalpha() and end.isalpha():
        first, last = ord(start), ord(end)
        _guard_axis_cardinality(first, last, step)
        return [chr(number) for number in range(first, last + 1, step)]
    raise ValidationError(
        "El rango debe ser numérico, alfabético simple o compartir un prefijo numérico.",
        code="location_axis_range_unsupported",
    )


def _canonical_import_headers(fieldnames: Sequence[object]) -> dict[str, object]:
    canonical: dict[str, object] = {}
    for original in fieldnames:
        target = _HEADER_ALIASES.get(_header(original))
        if target is None:
            continue
        if target in canonical:
            raise ValidationError(
                f"Más de una columna representa '{target}'.",
                code="location_import_header_repeated",
            )
        canonical[target] = original
    if not canonical or not {"aisle", "rack", "level", "position"}.issubset(canonical):
        raise ValidationError(
            "El archivo debe reconocer pasillo, rack, nivel y posición.",
            code="location_import_headers_invalid",
        )
    return canonical


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValidationError(
            "El CSV debe estar codificado en UTF-8.", code="location_import_encoding_invalid"
        ) from exc
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    previous_limit = csv.field_size_limit()
    csv.field_size_limit(1_000_000)
    try:
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        if not reader.fieldnames:
            raise ValidationError(
                "El archivo no contiene encabezados.", code="location_import_headers_missing"
            )
        canonical = _canonical_import_headers(reader.fieldnames)
        rows: list[dict[str, Any]] = []
        for raw in reader:
            if any(_nfkc_text(value) for value in raw.values()):
                rows.append({target: raw.get(original) for target, original in canonical.items()})
            if len(rows) > MAX_BATCH_ROWS:
                raise ValidationError(
                    f"El archivo contiene más de {MAX_BATCH_ROWS} filas.",
                    code="location_batch_too_large",
                )
    except csv.Error as exc:
        raise ValidationError(
            "El CSV contiene un campo demasiado grande o una estructura inválida.",
            code="location_import_field_too_large",
        ) from exc
    finally:
        csv.field_size_limit(previous_limit)
    return rows


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:  # noqa: C901
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise ValidationError(
            "El soporte XLSX no está disponible en el servidor.",
            code="location_import_xlsx_unavailable",
        ) from exc
    if not content.startswith(b"PK"):
        raise ValidationError("El archivo XLSX no es válido.", code="location_import_xlsx_invalid")
    _validate_xlsx_archive(content)
    try:
        workbook = load_workbook(
            io.BytesIO(content), read_only=True, data_only=True, keep_links=False
        )
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = next(iterator, None)
        if not headers:
            _fail_validation(
                "El archivo no contiene encabezados.", code="location_import_headers_missing"
            )
        canonical = [_HEADER_ALIASES.get(_header(value)) for value in headers]
        recognized = [target for target in canonical if target]
        if len(recognized) != len(set(recognized)):
            _fail_validation(
                "El XLSX contiene encabezados equivalentes repetidos.",
                code="location_import_header_repeated",
            )
        if not {"aisle", "rack", "level", "position"}.issubset(recognized):
            _fail_validation(
                "El archivo debe reconocer pasillo, rack, nivel y posición.",
                code="location_import_headers_invalid",
            )
        rows: list[dict[str, Any]] = []
        for values in iterator:
            if not any(_nfkc_text(value) for value in values):
                continue
            rows.append(
                {
                    target: values[index]
                    for index, target in enumerate(canonical)
                    if target is not None and index < len(values)
                }
            )
            if len(rows) > MAX_BATCH_ROWS:
                _fail_validation(
                    f"El archivo contiene más de {MAX_BATCH_ROWS} filas.",
                    code="location_batch_too_large",
                )
    except ValidationError:
        raise
    except Exception as exc:
        raise ValidationError(
            "El archivo XLSX no es válido o está dañado.", code="location_import_xlsx_invalid"
        ) from exc
    finally:
        if "workbook" in locals():
            workbook.close()
    return rows


class LocationUseCases:
    def __init__(self, repository: LocationRepository) -> None:
        self._repository = repository

    async def warehouse_scope(self, warehouse_id: uuid.UUID) -> WarehouseLocationScope:
        return await self._repository.get_warehouse_scope(warehouse_id)

    async def batch_scope(self, job_id: uuid.UUID) -> WarehouseLocationScope:
        return await self._repository.get_batch_scope(job_id)

    async def get_scheme(
        self, warehouse_id: uuid.UUID, version: int | None = None
    ) -> LocationCodeScheme:
        return await self._repository.get_scheme(warehouse_id, version)

    async def update_scheme(
        self,
        warehouse_id: uuid.UUID,
        *,
        name: str,
        separator: str,
        segments: Sequence[Mapping[str, Any]],
        actor_id: uuid.UUID,
    ) -> LocationCodeScheme:
        validated = tuple(CodeSegment(**dict(segment)) for segment in segments)
        return await self._repository.create_scheme_version(
            warehouse_id,
            name=_nfkc_text(name),
            separator=_nfkc_text(separator),
            segments=[
                {
                    "key": item.key,
                    "label": item.label,
                    "prefix": item.prefix,
                    "width": item.width,
                    "pad_char": item.pad_char,
                    "required": item.required,
                }
                for item in validated
            ],
            actor_id=actor_id,
        )

    async def preview_code(
        self,
        warehouse_id: uuid.UUID,
        values: Mapping[str, Any],
        *,
        scheme_version: int | None = None,
        exclude_location_id: uuid.UUID | None = None,
    ) -> tuple[Any, bool, bool]:
        scheme = await self._repository.get_scheme(warehouse_id, scheme_version)
        _, projection = _normalize_physical_values(scheme, values)
        code_exists, coordinates_exist = await self._repository.check_projection_conflicts(
            warehouse_id,
            projection,
            exclude_location_id=exclude_location_id,
        )
        return projection, code_exists, coordinates_exist

    async def get_location(self, warehouse_id: uuid.UUID, location_id: uuid.UUID) -> LocationRecord:
        return await self._repository.get_location(warehouse_id, location_id)

    async def required_update_permissions(
        self,
        warehouse_id: uuid.UUID,
        location_id: uuid.UUID,
        values: Mapping[str, Any],
        *,
        scheme_version: int | None = None,
    ) -> tuple[str, ...]:
        current = await self._repository.get_location(warehouse_id, location_id, for_update=True)
        scheme = await self._repository.get_scheme(warehouse_id, scheme_version)
        normalized, _projection = _normalize_physical_values(scheme, values)
        required = {"locations.update"}
        physical = ("area", "aisle", "rack", "level", "position")
        if any((getattr(current, key) or "") != (normalized.get(key) or "") for key in physical):
            required.add("locations.recode")
        target_status = str(normalized["lifecycle_status"])
        target_active = bool(normalized["is_active"])
        if target_active and not current.is_active:
            required.add("locations.activate")
        elif not target_active and current.is_active:
            required.add("locations.deactivate")
        elif target_status != current.lifecycle_status:
            if target_status == "retired":
                required.add("locations.deactivate")
            else:
                required.add("locations.commission")
        return tuple(sorted(required))

    async def create_location(
        self,
        warehouse_id: uuid.UUID,
        values: Mapping[str, Any],
        *,
        actor_id: uuid.UUID,
        scheme_version: int | None = None,
    ) -> LocationRecord:
        scheme = await self._repository.get_scheme(warehouse_id, scheme_version)
        normalized, projection = _normalize_physical_values(scheme, values)
        normalized["code_source"] = "generated"
        return await self._repository.create_location(
            warehouse_id, projection=projection, values=normalized, actor_id=actor_id
        )

    async def update_location(
        self,
        warehouse_id: uuid.UUID,
        location_id: uuid.UUID,
        values: Mapping[str, Any],
        *,
        actor_id: uuid.UUID,
        scheme_version: int | None = None,
        expected_updated_at: datetime | None = None,
    ) -> LocationRecord:
        scheme = await self._repository.get_scheme(warehouse_id, scheme_version)
        normalized, projection = _normalize_physical_values(scheme, values)
        normalized["code_source"] = "generated"
        return await self._repository.update_location(
            warehouse_id,
            location_id,
            projection=projection,
            values=normalized,
            actor_id=actor_id,
            expected_updated_at=expected_updated_at,
        )

    async def list_locations(self, warehouse_id: uuid.UUID, **filters: Any) -> Any:
        return await self._repository.list_locations(warehouse_id, **filters)

    async def summary(self, warehouse_id: uuid.UUID) -> dict[str, Any]:
        return await self._repository.location_summary(warehouse_id)

    async def preview_generator(
        self,
        warehouse_id: uuid.UUID,
        *,
        axes: Sequence[Mapping[str, Any]],
        defaults: Mapping[str, Any],
        idempotency_key: str,
        actor_id: uuid.UUID,
        scheme_version: int | None = None,
    ) -> LocationBatchRecord:
        scheme = await self._repository.get_scheme(warehouse_id, scheme_version)
        keys = [_nfkc_text(axis.get("key"), upper=False).casefold() for axis in axes]
        if not keys or any(key not in LOCATION_COMPONENT_KEYS for key in keys):
            raise ValidationError(
                "Los ejes del generador no son válidos.", code="location_axes_invalid"
            )
        if len(keys) != len(set(keys)):
            raise ValidationError("No se puede repetir un eje.", code="location_axis_repeated")
        expanded = [_expand_axis(axis) for axis in axes]
        total = 1
        for values in expanded:
            total *= len(values)
        if total > MAX_BATCH_ROWS:
            raise ValidationError(
                f"El lote generaría {total} ubicaciones; el máximo es {MAX_BATCH_ROWS}.",
                code="location_batch_too_large",
            )
        raw_rows = [
            {**defaults, **dict(zip(keys, combination, strict=True))}
            for combination in itertools.product(*expanded)
        ]
        source_rows = self._prepare_batch_rows(scheme, raw_rows, source="generated")
        checksum_payload = {
            "scheme_version": scheme.version,
            "axes": [dict(axis) for axis in axes],
            "defaults": dict(defaults),
        }
        return await self._repository.create_batch_preview(
            warehouse_id,
            kind="generate",
            idempotency_key=idempotency_key,
            input_checksum=_canonical_checksum(checksum_payload),
            scheme=scheme,
            source_rows=source_rows,
            actor_id=actor_id,
        )

    async def preview_import(
        self,
        warehouse_id: uuid.UUID,
        *,
        filename: str,
        content: bytes,
        idempotency_key: str,
        actor_id: uuid.UUID,
        scheme_version: int | None = None,
    ) -> LocationBatchRecord:
        if not content or len(content) > MAX_IMPORT_BYTES:
            raise ValidationError(
                "El archivo debe pesar entre 1 byte y 20 MB.",
                code="location_import_size_invalid",
            )
        extension = filename.rsplit(".", 1)[-1].casefold() if "." in filename else ""
        if extension == "csv":
            raw_rows = _parse_csv(content)
        elif extension == "xlsx":
            raw_rows = _parse_xlsx(content)
        else:
            raise ValidationError(
                "Solo se admiten archivos CSV o XLSX.", code="location_import_type_invalid"
            )
        if not raw_rows:
            raise ValidationError(
                "El archivo no contiene filas de datos.", code="location_import_empty"
            )
        if len(raw_rows) > MAX_BATCH_ROWS:
            raise ValidationError(
                f"El archivo contiene más de {MAX_BATCH_ROWS} filas.",
                code="location_batch_too_large",
            )
        scheme = await self._repository.get_scheme(warehouse_id, scheme_version)
        source_rows = self._prepare_batch_rows(scheme, raw_rows, source="imported")
        return await self._repository.create_batch_preview(
            warehouse_id,
            kind="import",
            idempotency_key=idempotency_key,
            input_checksum=_canonical_checksum(
                {
                    "content_sha256": hashlib.sha256(content).hexdigest(),
                    "extension": extension,
                    "scheme_version": scheme.version,
                }
            ),
            scheme=scheme,
            source_rows=source_rows,
            actor_id=actor_id,
        )

    def _prepare_batch_rows(
        self,
        scheme: LocationCodeScheme,
        raw_rows: Sequence[Mapping[str, Any]],
        *,
        source: str,
    ) -> list[dict[str, Any]]:
        result: list[dict[str, Any]] = []
        for row_number, raw in enumerate(raw_rows, start=2):
            try:
                provided_fields = _present_import_fields(raw) if source == "imported" else []
                normalized, projection = _normalize_physical_values(scheme, raw)
                normalized.update(
                    {
                        "code": projection.code,
                        "code_source": source,
                        "scheme_id": str(projection.scheme_id),
                        "scheme_version": projection.scheme_version,
                        "_provided_fields": provided_fields,
                    }
                )
                result.append({"row_number": row_number, **normalized})
            except ValidationError as exc:
                result.append(
                    {
                        "row_number": row_number,
                        "_errors": [exc.message],
                        "_error_code": exc.code,
                        "raw": dict(raw),
                    }
                )
        return result

    async def get_batch(
        self, job_id: uuid.UUID, *, page: int = 1, size: int = 100
    ) -> LocationBatchRecord:
        return await self._repository.get_batch(job_id, page=page, size=size)

    async def batch_required_permissions(self, job_id: uuid.UUID) -> tuple[str, ...]:
        return await self._repository.batch_required_permissions(job_id)

    async def publish_batch(self, job_id: uuid.UUID, *, actor_id: uuid.UUID) -> LocationBatchRecord:
        job = await self._repository.get_batch(job_id)
        if job.conflict_count or job.error_count:
            raise ConflictError(
                "Corrija los conflictos y errores antes de publicar el lote.",
                code="location_batch_not_publishable",
            )
        return await self._repository.publish_batch(job_id, actor_id=actor_id)
